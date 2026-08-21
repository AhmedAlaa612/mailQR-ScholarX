"""
One-off repair for registrations posted by the old Apps Script.

The old script addressed the form response sheet by hard-coded column index,
using an earlier version of the form. On the current sheet that meant:

    index 4 -> national_id   is actually "Governorate"
    index 5 -> governorate   is actually "What are your goals for attending?"
    index 7 -> institution   is actually "What is your question for the Session"

so participants.city holds the goals answer, participants.affiliation holds
the session question, and the real governorate was discarded.

This script re-reads the responses sheet and writes the correct city and
affiliation back onto participants, matched by email. Emails are normalized
with main.clean_email, because that is what the API stored: a registrant who
typed "@gamil.com" is in the database as "@gmail.com", and matching on the raw
sheet value would skip them.

It deliberately does NOT go through /api/register: that endpoint sends a
ticket email and inserts a new event_participants row on every call, so
replaying the sheet through it would mail every registrant again.

Not part of the deployed API — run it locally. Needs openpyxl in addition
to the packages in requirements.txt:

    pip install openpyxl

Usage:
    python backfill_form_columns.py responses.xlsx            # dry run
    python backfill_form_columns.py responses.xlsx --apply    # write
"""

import argparse
import os
import sys

from dotenv import load_dotenv
from supabase import create_client

load_dotenv()

# Substring matched (case-insensitive) against the sheet's header row.
COLUMNS = {
    "email":       "email",
    "governorate": "governorate",
    "university":  "university name",
    "college":     "collage name",   # matches the form's spelling
    "highschool":  "high school",
}
REQUIRED = ("email", "governorate")


def normalize(value):
    return " ".join(str("" if value is None else value).split()).lower()


def resolve_columns(headers):
    normalized = [normalize(h) for h in headers]
    index = {}
    for field, needle in COLUMNS.items():
        found = next((i for i, h in enumerate(normalized) if needle in h), None)
        if found is None:
            if field in REQUIRED:
                sys.exit(
                    f'Cannot find the "{needle}" column. Headers: {" | ".join(map(str, headers))}'
                )
            print(f'  warning: optional column "{needle}" not found')
        else:
            index[field] = found
    return index


def read_sheet(path):
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
    if not rows:
        sys.exit("Sheet is empty.")
    return rows[0], rows[1:]


def cell(row, index):
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def build_updates(headers, rows, clean_email):
    """Return {email: {city, affiliation}}, last submission per email winning."""
    col = resolve_columns(headers)
    updates = {}
    for row in rows:
        raw_email = cell(row, col.get("email"))
        if not raw_email:
            continue
        # Match the normalization the API applied before inserting the row.
        email = clean_email(raw_email)

        university = cell(row, col.get("university"))
        college = cell(row, col.get("college"))
        highschool = cell(row, col.get("highschool"))
        if university:
            affiliation = f"{university} - {college}" if college else university
        else:
            affiliation = highschool

        record = {}
        governorate = cell(row, col.get("governorate"))
        if governorate:
            record["city"] = governorate
        if affiliation:
            record["affiliation"] = affiliation
        if record:
            updates[email] = record
    return updates


def run():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("sheet", help="Responses sheet exported as .xlsx")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write to the database (default is a dry run)",
    )
    args = parser.parse_args()

    url, key = os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY")
    if not url or not key:
        sys.exit("SUPABASE_URL and SUPABASE_KEY must be set.")

    # Reuse the API's own cleaner so the two can never drift apart.
    # Importing main requires the Supabase env vars checked just above.
    from main import clean_email

    headers, rows = read_sheet(args.sheet)
    updates = build_updates(headers, rows, clean_email)
    print(f"{len(rows)} responses -> {len(updates)} unique emails")

    supabase = create_client(url, key)
    changed = missing = unchanged = 0

    for email, record in updates.items():
        found = (
            supabase.table("participants")
            .select("id, email, city, affiliation")
            .eq("email", email)
            .execute()
        )
        if not found.data:
            missing += 1
            print(f"  no participant row: {email}")
            continue

        participant = found.data[0]
        diff = {k: v for k, v in record.items() if participant.get(k) != v}
        if not diff:
            unchanged += 1
            continue

        changed += 1
        for field, value in diff.items():
            print(f"  {email}: {field}: {participant.get(field)!r} -> {value!r}")
        if args.apply:
            supabase.table("participants").update(diff).eq("id", participant["id"]).execute()

    verb = "updated" if args.apply else "would update"
    print(f"\n{verb} {changed}, already correct {unchanged}, not in database {missing}")
    if not args.apply:
        print("Dry run. Re-run with --apply to write.")


if __name__ == "__main__":
    run()
